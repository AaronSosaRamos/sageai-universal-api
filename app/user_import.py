"""
Plantilla e importación masiva de usuarios desde Excel (.xlsx).
Columnas: nombre, apellido, email, password, user_type (opcional: user | admin).
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass
from email.utils import parseaddr
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().replace("á", "a").replace("é", "e").replace("í", "i").replace("ó", "o").replace("ú", "u")


def build_user_import_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "usuarios"
    headers = ["nombre", "apellido", "email", "password", "user_type"]
    ws.append(headers)
    ws.append(
        [
            "Ejemplo",
            "Alumno",
            "reemplazar@institucion.edu",
            "CambiarPass1",
            "user",
        ]
    )
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 22
    ws2 = wb.create_sheet("instrucciones")
    ws2.append(["Instrucciones"])
    ws2.append(
        [
            "1. Elimina o sustituye la fila de ejemplo.",
            "2. user_type: 'user' (por defecto) o 'admin'.",
            "3. Contraseña: mínimo 8 caracteres, una mayúscula, una minúscula y un número.",
            "4. Guarda como .xlsx y súbelo en la herramienta de importación (solo administradores).",
        ]
    )
    ws2.column_dimensions["A"].width = 90
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@dataclass
class ParsedRow:
    row_num: int
    nombre: str
    apellido: str
    email: str
    password: str
    user_type: str


def _validate_password(pw: str) -> Optional[str]:
    if len(pw) < 8 or len(pw) > 50:
        return "La contraseña debe tener entre 8 y 50 caracteres"
    if not re.search(r"[A-Z]", pw):
        return "La contraseña debe incluir al menos una mayúscula"
    if not re.search(r"[a-z]", pw):
        return "La contraseña debe incluir al menos una minúscula"
    if not re.search(r"[0-9]", pw):
        return "La contraseña debe incluir al menos un número"
    return None


def _validate_name(s: str, field: str) -> Optional[str]:
    if len(s) < 2 or len(s) > 50:
        return f"{field} debe tener entre 2 y 50 caracteres"
    if not re.match(r"^[a-zA-ZáéíóúÁÉÍÓÚñÑ\s]+$", s):
        return f"{field} solo puede contener letras y espacios"
    return None


def parse_user_import_xlsx(content: bytes) -> Tuple[List[ParsedRow], List[Dict[str, Any]]]:
    """
    Parsea el archivo. Retorna (filas válidas, errores de formato por fila).
    Errores: {"row": int, "email": str, "error": str}
    """
    buf = io.BytesIO(content)
    wb = load_workbook(buf, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        return [], [{"row": 0, "email": "", "error": "El archivo está vacío"}]

    col_map: Dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key in ("nombre", "apellido", "email", "password", "user_type"):
            col_map[key] = i

    required = ("nombre", "apellido", "email", "password")
    missing = [k for k in required if k not in col_map]
    if missing:
        return [], [
            {
                "row": 0,
                "email": "",
                "error": f"Faltan columnas obligatorias: {', '.join(missing)}. Usa la plantilla descargada.",
            }
        ]

    valid: List[ParsedRow] = []
    errors: List[Dict[str, Any]] = []
    row_num = 1

    for row in rows_iter:
        row_num += 1
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue

        def cell(name: str) -> str:
            idx = col_map[name]
            v = row[idx] if idx < len(row) else None
            if v is None:
                return ""
            return str(v).strip()

        nombre = cell("nombre")
        apellido = cell("apellido")
        email = cell("email").lower()
        password = cell("password")
        ut_raw = cell("user_type") if "user_type" in col_map else ""
        user_type = (ut_raw or "user").strip().lower()
        if user_type not in ("user", "admin"):
            user_type = "user"

        if not email:
            continue

        _, addr = parseaddr(email)
        if not addr or "@" not in addr:
            errors.append({"row": row_num, "email": email, "error": "Email inválido"})
            continue

        err = _validate_name(nombre, "Nombre")
        if err:
            errors.append({"row": row_num, "email": email, "error": err})
            continue
        err = _validate_name(apellido, "Apellido")
        if err:
            errors.append({"row": row_num, "email": email, "error": err})
            continue
        err = _validate_password(password)
        if err:
            errors.append({"row": row_num, "email": email, "error": err})
            continue

        valid.append(
            ParsedRow(
                row_num=row_num,
                nombre=nombre,
                apellido=apellido,
                email=addr,
                password=password,
                user_type=user_type,
            )
        )

    return valid, errors
